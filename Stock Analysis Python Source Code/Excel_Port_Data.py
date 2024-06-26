import requests
from bs4 import BeautifulSoup
from transformers import AutoTokenizer, AutoModelForSequenceClassification
import torch
import numpy as np
from datetime import datetime, timedelta
import pandas as pd
import yfinance as yf
from langdetect import detect
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def get_stock_codes_and_names():
    url = "https://site2.sbisec.co.jp/ETGate/?OutSide=on&_ControlID=WPLETmgR001Control&_PageID=WPLETmgR001Mdtl20&_DataStoreID=DSWPLETmgR001Control&_ActionID=DefaultAID&getFlg=on&burl=search_market&cat1=market&cat2=none&dir=info&file=market_meigara_225.html"
    
    response = requests.get(url)
    if response.status_code != 200:
        print(f"Failed to fetch the webpage. Status code: {response.status_code}")
        return []

    soup = BeautifulSoup(response.content, 'html.parser')
    stock_table = soup.find('table', {'class': 'md-l-table-type01'})
    if stock_table is None:
        all_tables = soup.find_all('table')
        for table in all_tables:
            if table.find('tr'):
                stock_table = table
                break
        if stock_table is None:
            return []

    stock_data = []
    for row in stock_table.find_all('tr')[1:]:
        cells = row.find_all('td')
        if cells:
            stock_code = cells[0].text.strip()
            company_name = cells[1].text.strip()
            stock_data.append((stock_code, company_name))
    return stock_data

def scrape_nikkei_news(stock_number):
    url = f"https://www.nikkei.com/nkd/company/news/?scode={stock_number}&ba=1"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    news_items = soup.find_all('a', href=lambda href: href and "/nkd/company/article/" in href)
    news_data = [{"title": item.text.strip(), "url": "https://www.nikkei.com" + item['href']} for item in news_items]
    return news_data

def scrape_yahoo_finance_news(stock_number):
    ticker = f"{stock_number}.T"
    url = f"https://finance.yahoo.co.jp/quote/{ticker}/news"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    news_items = soup.find_all('a', href=lambda href: href and "/news/" in href)
    news_data = []
    for item in news_items:
        title = item.text.strip()
        article_url = item['href']
        if not article_url.startswith('http'):
            article_url = "https://finance.yahoo.co.jp" + article_url
        news_data.append({"title": title, "url": article_url})
    return news_data

def analyze_sentiment(text, ja_tokenizer, ja_model, en_tokenizer, en_model):
    try:
        lang = detect(text)
    except:
        lang = 'ja'  # Default to Japanese if detection fails

    tokenizer = ja_tokenizer if lang == 'ja' else en_tokenizer
    model = ja_model if lang == 'ja' else en_model

    inputs = tokenizer(text, return_tensors="pt", truncation=True, max_length=512)
    outputs = model(**inputs)
    sentiment_score = torch.softmax(outputs.logits, dim=1).tolist()[0]
    return sentiment_score[0]  # Return the raw sentiment score

def sentiment_to_text(score):
    if score > 0.8:
        return "Very Negative"
    elif score > 0.6:
        return "Negative"
    elif score > 0.4:
        return "Neutral"
    elif score > 0.2:
        return "Positive"
    else:
        return "Very Positive"

def calculate_average_sentiment(sentiments):
    if not sentiments:
        return "Neutral"
    avg_sentiment = sum(sentiments) / len(sentiments)
    return sentiment_to_text(avg_sentiment)

def get_stock_data(stock_number):
    ticker = f"{stock_number}.T"
    end_date = datetime.now()
    start_date = end_date - timedelta(days=30)
    
    try:
        df = yf.download(ticker, start=start_date, end=end_date)
        if df.empty:
            return None
        
        df = df.reset_index()
        df['Date'] = pd.to_datetime(df['Date'])
        stock_data = [(row['Date'], row['Close']) for _, row in df.iterrows()]
        stock_data.sort(key=lambda x: x[0], reverse=True)
        return stock_data[:30]
    
    except Exception as e:
        print(f"Error retrieving stock data: {e}")
        return None

def calculate_stock_trend(stock_data):
    if not stock_data or len(stock_data) < 2:
        return "No trend data"
    
    first_price = stock_data[-1][1]  # Oldest price
    last_price = stock_data[0][1]   # Most recent price
    
    percent_change = ((last_price - first_price) / first_price) * 100
    
    if percent_change > 5:
        return "Strong Uptrend"
    elif percent_change > 2:
        return "Uptrend"
    elif percent_change < -5:
        return "Strong Downtrend"
    elif percent_change < -2:
        return "Downtrend"
    else:
        return "Neutral"

def get_action_recommendation(public_opinion, stock_trend, stock_price_data, purchase_price=None):
    if not stock_price_data:
        return "Insufficient data for recommendation"

    opinion_score = {"Very Positive": 2, "Positive": 1, "Neutral": 0, "Negative": -1, "Very Negative": -2}
    trend_score = {"Strong Uptrend": 2, "Uptrend": 1, "Neutral": 0, "Downtrend": -1, "Strong Downtrend": -2}
    
    total_score = opinion_score.get(public_opinion, 0) + trend_score.get(stock_trend, 0)
    
    prices = [price for _, price in stock_price_data]
    current_price = prices[0]
    avg_price = np.mean(prices)
    std_dev = np.std(prices)
    
    owns_stock = purchase_price is not None
    
    if owns_stock:
        price_change = (current_price - purchase_price) / purchase_price * 100
        
        if total_score > 0:
            action = "Hold"
            explanation = f"Positive outlook. You're currently up {price_change:.2f}%. Consider holding for potential further gains."
        elif total_score < 0:
            action = "Consider Selling"
            explanation = f"Negative outlook. You're currently {'up' if price_change > 0 else 'down'} {abs(price_change):.2f}%. Consider selling to {'lock in profits' if price_change > 0 else 'minimize losses'}."
        else:
            action = "Hold and Monitor"
            explanation = f"Mixed signals. You're currently {'up' if price_change > 0 else 'down'} {abs(price_change):.2f}%. Monitor the stock closely for changes in sentiment or market trends."
        
        if price_change > 20:
            explanation += " However, with significant gains, consider taking partial profits."
        elif price_change < -20:
            explanation += " However, with significant losses, reassess your investment thesis."
    else:
        if total_score > 0:
            target_price = max(current_price * 0.99, avg_price - 0.5 * std_dev)
            action = f"Consider Buying (Target: ¥{target_price:.2f})"
            explanation = "Positive outlook. Consider buying near the suggested target price."
        elif total_score < 0:
            action = "Hold Off"
            explanation = "Negative outlook. It might be better to wait for a more favorable entry point."
        else:
            action = "Monitor"
            explanation = "Mixed signals. Monitor the stock for a clearer trend before making a decision."

    return f"{action}\nExplanation: {explanation}"

def get_yahoo_finance_price(stock_number):
    url = f"https://finance.yahoo.co.jp/quote/{stock_number}.T"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    price_element = soup.find('span', class_='_3rXWJKZF')
    if price_element:
        price_text = price_element.text.replace(',', '')
        try:
            return int(float(price_text))
        except ValueError:
            print(f"Warning: Unable to convert price to integer for stock {stock_number}: {price_text}")
            return None
    return None

def create_excel_report(stock_analysis):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Analysis"

    # Set up header row (company names)
    for col, stock in enumerate(stock_analysis, start=2):
        cell = ws.cell(row=1, column=col, value=stock['company_name'])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Set up row headers
    row_headers = [
        "Stock Code",
        f"{datetime.now().date()} Stock Price",
        "Previous Day Stock Price",
        "Compared to day before",
        "Nikkei Perception",
        "Yahoo Finance Perception",
        "Overall Perception",
        "Action"
    ]
    for row, header in enumerate(row_headers, start=2):
        cell = ws.cell(row=row, column=1, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[row].height = 20

    def color_sentiment(cell, sentiment):
        if sentiment == "Positive" or sentiment == "Very Positive":
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif sentiment == "Negative" or sentiment == "Very Negative":
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif sentiment == "Neutral":
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Populate data
    for col, stock in enumerate(stock_analysis, start=2):
        current_price = stock['current_stock_price']
        previous_price = stock['stock_price_data'][1][1] if len(stock['stock_price_data']) > 1 else None
        
        if current_price is not None and previous_price is not None:
            price_change_percent = ((current_price - previous_price) / previous_price) * 100
        else:
            price_change_percent = "N/A"

        ws.cell(row=2, column=col, value=stock['stock_number'])
        ws.cell(row=3, column=col, value=current_price if current_price is not None else "N/A")
        ws.cell(row=4, column=col, value=previous_price if previous_price is not None else "N/A")
        
        change_cell = ws.cell(row=5, column=col, value=f"{price_change_percent:.2f}%" if price_change_percent != "N/A" else price_change_percent)
        if price_change_percent != "N/A":
            if price_change_percent > 0:
                change_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
            elif price_change_percent < 0:
                change_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            else:
                change_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        nikkei_cell = ws.cell(row=6, column=col, value=stock['nikkei_sentiment'])
        color_sentiment(nikkei_cell, stock['nikkei_sentiment'])
        
        yahoo_cell = ws.cell(row=7, column=col, value=stock['yahoo_sentiment'])
        color_sentiment(yahoo_cell, stock['yahoo_sentiment'])
        
        overall_cell = ws.cell(row=8, column=col, value=stock['overall_sentiment'])
        color_sentiment(overall_cell, stock['overall_sentiment'])
        
        action_cell = ws.cell(row=9, column=col, value=stock['action_recommendation'].split('\n')[0])  # Only the action, not the explanation
        if "Buy" in action_cell.value:
            action_cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Save the workbook
    filename = f"stock_analysis_{datetime.now().strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    print(f"Excel report saved as {filename}")

def main():
    stock_data = get_stock_codes_and_names()
    if not stock_data:
        print("Failed to retrieve stock data. Please check your internet connection and try again.")
        return

    total_stocks = len(stock_data)
    print(f"Total stocks to process: {total_stocks}")

    ja_tokenizer = AutoTokenizer.from_pretrained("jarvisx17/japanese-sentiment-analysis")
    ja_model = AutoModelForSequenceClassification.from_pretrained("jarvisx17/japanese-sentiment-analysis")
    
    en_tokenizer = AutoTokenizer.from_pretrained("distilbert-base-uncased-finetuned-sst-2-english")
    en_model = AutoModelForSequenceClassification.from_pretrained("distilbert-base-uncased-finetuned-sst-2-english")

    stock_analysis = []

    for index, (stock_number, company_name) in enumerate(stock_data, 1):
        try:
            print(f"Processing: {company_name} ({stock_number}) - {index}/{total_stocks}", end='\r')
            
            nikkei_news_data = scrape_nikkei_news(stock_number)
            yahoo_finance_news_data = scrape_yahoo_finance_news(stock_number)
            
            nikkei_sentiments = [analyze_sentiment(news['title'], ja_tokenizer, ja_model, en_tokenizer, en_model) for news in nikkei_news_data]
            yahoo_finance_sentiments = [analyze_sentiment(news['title'], ja_tokenizer, ja_model, en_tokenizer, en_model) for news in yahoo_finance_news_data]
            
            nikkei_overall_sentiment = calculate_average_sentiment(nikkei_sentiments)
            yahoo_finance_overall_sentiment = calculate_average_sentiment(yahoo_finance_sentiments)
            
            overall_sentiment_value = (sum(nikkei_sentiments) + sum(yahoo_finance_sentiments)) / (len(nikkei_sentiments) + len(yahoo_finance_sentiments)) if nikkei_sentiments or yahoo_finance_sentiments else 0.5
            overall_sentiment = sentiment_to_text(overall_sentiment_value)
            
            stock_price_data = get_stock_data(stock_number)
            current_stock_price = get_yahoo_finance_price(stock_number)
            stock_trend = calculate_stock_trend(stock_price_data)
            
            action_recommendation = get_action_recommendation(overall_sentiment, stock_trend, stock_price_data)
            
            stock_analysis.append({
                'stock_number': stock_number,
                'company_name': company_name,
                'current_stock_price': current_stock_price,
                'nikkei_sentiment': nikkei_overall_sentiment,
                'yahoo_sentiment': yahoo_finance_overall_sentiment,
                'overall_sentiment': overall_sentiment,
                'stock_trend': stock_trend,
                'action_recommendation': action_recommendation,
                'stock_price_data': stock_price_data
            })
        
        except Exception as e:
            print(f"\nError processing {company_name} ({stock_number}): {str(e)}")

    print("\nAll stocks processed.")
    create_excel_report(stock_analysis)

if __name__ == '__main__':
    main()