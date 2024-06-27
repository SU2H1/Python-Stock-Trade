import requests
from bs4 import BeautifulSoup
from transformers import AutoTokenizer, AutoModelForSequenceClassification
import torch
import numpy as np
from datetime import datetime, timedelta
import pandas as pd
import yfinance as yf
from langdetect import detect
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os

def get_stock_codes_and_names():
    url = "https://site2.sbisec.co.jp/ETGate/?OutSide=on&_ControlID=WPLETmgR001Control&_PageID=WPLETmgR001Mdtl20&_DataStoreID=DSWPLETmgR001Control&_ActionID=DefaultAID&getFlg=on&burl=search_market&cat1=market&cat2=none&dir=info&file=market_meigara_225.html"
    
    response = requests.get(url)
    if response.status_code != 200:
        print(f"Failed to fetch the webpage. Status code: {response.status_code}")
        return []

    soup = BeautifulSoup(response.content, 'html.parser')
    stock_table = soup.find('table', {'class': 'md-l-table-type01'})
    if stock_table is None:
        all_tables = soup.find_all('table')
        for table in all_tables:
            if table.find('tr'):
                stock_table = table
                break
        if stock_table is None:
            return []

    stock_data = []
    for row in stock_table.find_all('tr')[1:]:
        cells = row.find_all('td')
        if cells:
            stock_code = cells[0].text.strip()
            company_name = cells[1].text.strip()
            stock_data.append((stock_code, company_name))
    return stock_data

def scrape_nikkei_news(stock_number):
    url = f"https://www.nikkei.com/nkd/company/news/?scode={stock_number}&ba=1"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    news_items = soup.find_all('a', href=lambda href: href and "/nkd/company/article/" in href)
    news_data = [{"title": item.text.strip(), "url": "https://www.nikkei.com" + item['href']} for item in news_items]
    return news_data

def scrape_yahoo_finance_news(stock_number):
    ticker = f"{stock_number}.T"
    url = f"https://finance.yahoo.co.jp/quote/{ticker}/news"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    news_items = soup.find_all('a', href=lambda href: href and "/news/" in href)
    news_data = []
    for item in news_items:
        title = item.text.strip()
        article_url = item['href']
        if not article_url.startswith('http'):
            article_url = "https://finance.yahoo.co.jp" + article_url
        news_data.append({"title": title, "url": article_url})
    return news_data

def analyze_sentiment(text, ja_tokenizer, ja_model, en_tokenizer, en_model):
    try:
        lang = detect(text)
    except:
        lang = 'ja'  # Default to Japanese if detection fails

    tokenizer = ja_tokenizer if lang == 'ja' else en_tokenizer
    model = ja_model if lang == 'ja' else en_model

    inputs = tokenizer(text, return_tensors="pt", truncation=True, max_length=512)
    outputs = model(**inputs)
    sentiment_score = torch.softmax(outputs.logits, dim=1).tolist()[0]
    return sentiment_score[0]  # Return the raw sentiment score

def sentiment_to_text(score):
    if score > 0.8:
        return "Very Negative"
    elif score > 0.6:
        return "Negative"
    elif score > 0.4:
        return "Neutral"
    elif score > 0.2:
        return "Positive"
    else:
        return "Very Positive"

def calculate_average_sentiment(sentiments):
    if not sentiments:
        return "Neutral"
    avg_sentiment = sum(sentiments) / len(sentiments)
    return sentiment_to_text(avg_sentiment)

def get_stock_data(stock_number):
    ticker = f"{stock_number}.T"
    end_date = datetime.now()
    start_date = end_date - timedelta(days=30)
    
    try:
        df = yf.download(ticker, start=start_date, end=end_date)
        if df.empty:
            return None
        
        df = df.reset_index()
        df['Date'] = pd.to_datetime(df['Date'])
        stock_data = [(row['Date'], row['Close']) for _, row in df.iterrows()]
        stock_data.sort(key=lambda x: x[0], reverse=True)
        return stock_data[:30]
    
    except Exception as e:
        print(f"Error retrieving stock data: {e}")
        return None

def calculate_stock_trend(stock_data):
    if not stock_data or len(stock_data) < 2:
        return "No trend data"
    
    first_price = stock_data[-1][1]  # Oldest price
    last_price = stock_data[0][1]   # Most recent price
    
    percent_change = ((last_price - first_price) / first_price) * 100
    
    if percent_change > 5:
        return "Strong Uptrend"
    elif percent_change > 2:
        return "Uptrend"
    elif percent_change < -5:
        return "Strong Downtrend"
    elif percent_change < -2:
        return "Downtrend"
    else:
        return "Neutral"

def get_action_recommendation(public_opinion, stock_trend, stock_price_data, purchase_price=None):
    if not stock_price_data:
        return "Insufficient data for recommendation"

    opinion_score = {"Very Positive": 2, "Positive": 1, "Neutral": 0, "Negative": -1, "Very Negative": -2}
    trend_score = {"Strong Uptrend": 2, "Uptrend": 1, "Neutral": 0, "Downtrend": -1, "Strong Downtrend": -2}
    
    total_score = opinion_score.get(public_opinion, 0) + trend_score.get(stock_trend, 0)
    
    prices = [price for _, price in stock_price_data]
    current_price = prices[0]
    avg_price = np.mean(prices)
    std_dev = np.std(prices)
    
    owns_stock = purchase_price is not None
    
    if owns_stock:
        price_change = (current_price - purchase_price) / purchase_price * 100
        
        if total_score > 0:
            action = "Hold"
            explanation = f"Positive outlook. You're currently up {price_change:.2f}%. Consider holding for potential further gains."
        elif total_score < 0:
            action = "Consider Selling"
            explanation = f"Negative outlook. You're currently {'up' if price_change > 0 else 'down'} {abs(price_change):.2f}%. Consider selling to {'lock in profits' if price_change > 0 else 'minimize losses'}."
        else:
            action = "Hold and Monitor"
            explanation = f"Mixed signals. You're currently {'up' if price_change > 0 else 'down'} {abs(price_change):.2f}%. Monitor the stock closely for changes in sentiment or market trends."
        
        if price_change > 20:
            explanation += " However, with significant gains, consider taking partial profits."
        elif price_change < -20:
            explanation += " However, with significant losses, reassess your investment thesis."
    else:
        if total_score > 0:
            target_price = max(current_price * 0.99, avg_price - 0.5 * std_dev)
            action = f"Consider Buying (Target: ¥{target_price:.2f})"
            explanation = "Positive outlook. Consider buying near the suggested target price."
        elif total_score < 0:
            action = "Hold Off"
            explanation = "Negative outlook. It might be better to wait for a more favorable entry point."
        else:
            action = "Monitor"
            explanation = "Mixed signals. Monitor the stock for a clearer trend before making a decision."

    return f"{action}\nExplanation: {explanation}"

def get_yahoo_finance_price(stock_number):
    url = f"https://finance.yahoo.co.jp/quote/{stock_number}.T"
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    price_element = soup.find('span', class_='_3rXWJKZF')
    if price_element:
        price_text = price_element.text.replace(',', '')
        try:
            return int(float(price_text))
        except ValueError:
            print(f"Warning: Unable to convert price to integer for stock {stock_number}: {price_text}")
            return None
    return None

def create_excel_report(stock_analysis):
    today = datetime.now().date()
    directory = r"C:\Users\ka1t0\Documents\Python-Stock-Trade"
    filename = os.path.join(directory, "Nikkei225_Stock_Analysis.xlsx")
    
    try:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Stock Analysis"

    # Set up headers if it's a new file
    if ws.max_column == 1:
        headers = ["Metric"] + [stock['company_name'] for stock in stock_analysis]
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)

    # Find the next empty row
    next_row = ws.max_row + 1

    # Define the metrics
    metrics = [
        f"{today} Stock Price",
        "Previous Day Stock Price",
        "Previous Day Stock Price Percentage",
        "Nikkei's Perception",
        "Yahoo Finance's Perception",
        "Overall Perception",
        "Recommended Action"
    ]

    def color_cell(cell, value):
        if value in ["Positive", "Very Positive"] or (isinstance(value, (float, int)) and value > 0):
            cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        elif value in ["Negative", "Very Negative"] or (isinstance(value, (float, int)) and value < 0):
            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
        elif value == "Neutral" or (isinstance(value, (float, int)) and value == 0):
            cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # Add new data
    for i, metric in enumerate(metrics):
        ws.cell(row=next_row + i, column=1, value=metric)
        for col, stock in enumerate(stock_analysis, start=2):
            if i == 0:  # Current day's stock price
                ws.cell(row=next_row + i, column=col, value=stock['current_stock_price'])
            elif i == 1:  # Previous day's stock price
                ws.cell(row=next_row + i, column=col, value=stock['previous_stock_price'])
            elif i == 2:  # Previous day's stock price percentage
                if stock['current_stock_price'] and stock['previous_stock_price']:
                    price_change = ((stock['current_stock_price'] - stock['previous_stock_price']) / stock['previous_stock_price']) * 100
                    change_cell = ws.cell(row=next_row + i, column=col, value=f"{price_change:.2f}%")
                    color_cell(change_cell, price_change)
                else:
                    ws.cell(row=next_row + i, column=col, value="N/A")
            elif i == 3:  # Nikkei's Perception
                cell = ws.cell(row=next_row + i, column=col, value=stock['nikkei_sentiment'])
                color_cell(cell, stock['nikkei_sentiment'])
            elif i == 4:  # Yahoo Finance's Perception
                cell = ws.cell(row=next_row + i, column=col, value=stock['yahoo_sentiment'])
                color_cell(cell, stock['yahoo_sentiment'])
            elif i == 5:  # Overall Perception
                cell = ws.cell(row=next_row + i, column=col, value=stock['overall_sentiment'])
                color_cell(cell, stock['overall_sentiment'])
            elif i == 6:  # Recommended Action
                cell = ws.cell(row=next_row + i, column=col, value=stock['action_recommendation'].split('\n')[0])
                if "Buy" in cell.value:
                    cell.fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

    # Adjust column widths
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    # Save the workbook
    attempt = 0
    while attempt < 5:  # Try up to 5 times
        try:
            if attempt == 0:
                save_filename = filename
            else:
                save_filename = os.path.join(directory, f"Nikkei225_Stock_Analysis_{attempt}.xlsx")
            wb.save(save_filename)
            print(f"Excel report updated and saved as {save_filename}")
            break
        except PermissionError:
            print(f"Unable to save to {save_filename}. Trying a different filename...")
            attempt += 1
    else:
        print("Failed to save the Excel file after multiple attempts. Please close any open instances of the file and try again.")

def main():
    stock_data = get_stock_codes_and_names()
    if not stock_data:
        print("Failed to retrieve stock data. Please check your internet connection and try again.")
        return

    total_stocks = len(stock_data)
    print(f"Total stocks to process: {total_stocks}")

    ja_tokenizer = AutoTokenizer.from_pretrained("jarvisx17/japanese-sentiment-analysis")
    ja_model = AutoModelForSequenceClassification.from_pretrained("jarvisx17/japanese-sentiment-analysis")
    
    en_tokenizer = AutoTokenizer.from_pretrained("distilbert-base-uncased-finetuned-sst-2-english")
    en_model = AutoModelForSequenceClassification.from_pretrained("distilbert-base-uncased-finetuned-sst-2-english")

    stock_analysis = []

    for index, (stock_number, company_name) in enumerate(stock_data, 1):
        try:
            print(f"Processing: {company_name} ({stock_number}) - {index}/{total_stocks}", end='\r')
            
            nikkei_news_data = scrape_nikkei_news(stock_number)
            yahoo_finance_news_data = scrape_yahoo_finance_news(stock_number)
            
            nikkei_sentiments = [analyze_sentiment(news['title'], ja_tokenizer, ja_model, en_tokenizer, en_model) for news in nikkei_news_data]
            yahoo_finance_sentiments = [analyze_sentiment(news['title'], ja_tokenizer, ja_model, en_tokenizer, en_model) for news in yahoo_finance_news_data]
            
            nikkei_overall_sentiment = calculate_average_sentiment(nikkei_sentiments)
            yahoo_finance_overall_sentiment = calculate_average_sentiment(yahoo_finance_sentiments)
            
            overall_sentiment_value = (sum(nikkei_sentiments) + sum(yahoo_finance_sentiments)) / (len(nikkei_sentiments) + len(yahoo_finance_sentiments)) if nikkei_sentiments or yahoo_finance_sentiments else 0.5
            overall_sentiment = sentiment_to_text(overall_sentiment_value)
            
            stock_price_data = get_stock_data(stock_number)
            current_stock_price = get_yahoo_finance_price(stock_number)
            previous_stock_price = stock_price_data[1][1] if len(stock_price_data) > 1 else None
            stock_trend = calculate_stock_trend(stock_price_data)
            
            action_recommendation = get_action_recommendation(overall_sentiment, stock_trend, stock_price_data)
            
            stock_analysis.append({
                'stock_number': stock_number,
                'company_name': company_name,
                'current_stock_price': current_stock_price,
                'previous_stock_price': previous_stock_price,
                'nikkei_sentiment': nikkei_overall_sentiment,
                'yahoo_sentiment': yahoo_finance_overall_sentiment,
                'overall_sentiment': overall_sentiment,
                'stock_trend': stock_trend,
                'action_recommendation': action_recommendation,
                'stock_price_data': stock_price_data
            })
        
        except Exception as e:
            print(f"\nError processing {company_name} ({stock_number}): {str(e)}")

    print("\nAll stocks processed.")
    create_excel_report(stock_analysis)

if __name__ == '__main__':
    main()
